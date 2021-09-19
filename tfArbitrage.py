#Dependencies. A whole lot of them.
import os
import re
import sys
import time
import json
import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import steampy
from steampy.guard import generate_one_time_code
from steampy.client import SteamClient
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import selenium
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


#Global Variables
#User Agent Variables
sheetDirectory = ".\tfArbitrage.xlsx"

#SteamPy Variables
#Read info.txt
infoDirectory = ".\info.txt"
if(os.path.exists(infoDirectory) != True):
    infoCreate()

#Load variables from info file
infoFile = open(".\info.txt", "r")
infoArray = infoFile.readlines()
if len(infoArray) < 10:
    print("info.txt file not filled out properly.")
    sys.exit()
# Set API
api_key = infoArray[1][:-1]
# Steam steamname
steamname = str(infoArray[3][:-1])
# Steam password
password = str(infoArray[5][:-1])
# steam's Shared Secret
secret = infoArray[7]
delay = 5
#Compile to find any number of digits, a decimal, then 1 to two digits
twoDecRegex = re.compile(r"\d*[.]\d{0,2}")

#Chunk to deal with bpTF API
#Get key from info file, remove the \n at the end of the line from using readlines method
apiKey = infoArray [9][:-1]
#Call the API to get the info
apiGet = 'https://backpack.tf/api/IGetPrices/v4?raw=1&key=' + apiKey
#Use requests to get the raw response
bpRequest = requests.get(apiGet).text
#Use JSON library to parse information so python can read it as a dict
bpResponse = json.loads(bpRequest)

#Get key to refined price from API, will use this to convert values later
keyToRef = float(bpResponse["response"]["items"]["Mann Co. Supply Crate Key"]["prices"]["6"]["Tradable"]["Craftable"][0]["value"])



#Function definitions
#Creates info.txt file
def infoCreate():
    infoFile = open(infoDirectory, "w+")
    infoFile.write('''#Go to ( https://steamcommunity.com/dev/apikey ) and paste API key below

#Put Username on line below

#Put Password on line below

#Put shared secret on line below

#Put Backpack.tf API Key below. Go to https://backpack.tf/developer/apikey/view for help.
 ''')
    infoFile.close()
    print('--File created at ' + infoDirectory)
    incompleteInfo()

def getAuthCode():
    steamPyAuthCode = generate_one_time_code(secret)
    return steamPyAuthCode

    #Logs into steam, given correct authcode

def scrapeScrap(authCode):

    #Browser setup
    browser = webdriver.Chrome(ChromeDriverManager().install())
    browser.maximize_window()
    browser.get("https://scrap.tf/buy/hats")
    wait = WebDriverWait(browser, delay)
    wait.until(EC.element_to_be_clickable((By.ID, "steamAccountName"))).send_keys(steamname)
    wait.until(EC.element_to_be_clickable((By.ID, "steamPassword"))).send_keys(password)
    wait.until(EC.element_to_be_clickable((By.ID, "steamPassword"))).send_keys(Keys.RETURN)
    wait.until(EC.element_to_be_clickable((By.ID, "twofactorcode_entry"))).send_keys(authCode)
    wait.until(EC.element_to_be_clickable((By.ID, "twofactorcode_entry"))).send_keys(Keys.RETURN)

    #Variable declaration
    elements = []
    file = 'tfArbitrage.xlsx'
    wb = load_workbook(filename=file)
    ws = wb.active
    refRegex = re.compile(r"\d[\d\s.,]*refined")
    keyRegex = re.compile(r"\d[\d\s.,]*keys")

    #Scraping the site for items it has available
    time.sleep(delay)
    #Use soup on finished page
    browser.get("https://scrap.tf/buy/hats")
    soup = BeautifulSoup(browser.page_source, "html.parser")
    #for thing in soup.find_all('div', class_=):
    #    print(thing)
    for container in soup.find_all('div', class_="items-container"):#Tracks down all divs in the results
        for element in container.find_all('div'):
            if element.get('data-appid') == "440":
                itemCost = element.get('data-content')
                dataID = element.get('data-id')#Only processes divs with a data-id attribute
                #############################################
                elemData = [] #Refresh elemData variable, to store information on each specific item

                #item ID Number codeblock
                #TODO: is the ID scraptf or tf2 ID number in general?
                elemData.append(dataID)#Item ID Number

                #Item name codeblock
                nameSoup = element.get('data-title')#Item Name. Sometimes the name is nested in an extra span tag.
                #If there's no name, no point in continuing this loop.
                if nameSoup == None:
                    continue

                while nameSoup[0] == "<":#This while loop iterates to make sure to pull the name out if it exists
                    nameSoup = BeautifulSoup(element.get('data-title'), "html.parser")
                    nameSoup = nameSoup.find('span').string

                elemData.append(nameSoup)#Item Name, after being verified.

                #Item Quality Number
                itemQual =  element['class'][2][7:]
                elemData.append(itemQual)

                #Item quantity codeblock
                itemQuantity = element.get('data-num-available')
                if itemQuantity == None:
                    #Check if item is available
                    #Break out of check if it isnt
                    continue
                elemData.append(itemQuantity)#Num of item available

                #Item price codeblock
                itemCost = element.get('data-content')
                #Using regex to grab html up to mention of "Keys"
                itemKey = keyRegex.search(itemCost)
                #Check to see if there is a key price. Append a blank string if there is none.
                if itemKey != None:
                    itemKey = itemKey.group(0)[:-4]

                #Use regex to grab html after keys to mention of refined
                itemRef = refRegex.search(itemCost)
                if itemRef == None:
                    continue
                else:
                    itemRef = itemRef.group(0)[:-7]

                #Converted price Codeblock
                #Check if there was a key price
                if itemKey != None:
                    itemRef = keyConvert(float(itemKey), float(itemRef))

                elemData.append(itemRef)


                #bptf price codeblock
                #Check if the start of the scraptf name is strange
                if itemQual == "11" or itemQual == "3":
                    #Cut it off if it is
                    nameSoup = nameSoup[8:]
                itemBP = bpResponse["response"]["items"][nameSoup]["prices"][itemQual]["Tradable"]["Craftable"][0]
                if itemBP["currency"] == "keys":
                    itemBP = keyConvert(float(itemBP["value"]), 0)
                else:
                    itemBP = itemBP["value"]

                elemData.append(itemBP)

                elemData.append(str(float(itemBP) - float(itemRef)))



                #Attaches each array of info to the 2d array of all information grabbed
                if(elemData[4] != ""):
                    elements.append(elemData)
                ########################################################



    #Write array onto spreadsheet
    for i in range(0, len(elements)):
        for j in range(0, len(elements[i])):
          ws.cell(row=i+2, column=j+1).value = str(elements[i][j])#Add 1 to i, so it has space to work with

    wb.save(file)

def keyConvert(keyFloat, refFloat):
    #Calculate value.
    #Each float gets multiplied by 9, then floored to an integer.
    #2 is added to compensate,
    #so .11 ref turns into .99, 0, then 1.
    #Additional multiplier given from the keytoref pulled from the api
    #This is how much the value is in scrap. (9 scrap to a ref)
    value = (int(keyFloat * 9 * keyToRef) + int(refFloat * 9) + 2)
    #Divide the value by 9 again
    value /= 9
    #Use a regex to only display two decimal places.
    value = twoDecRegex.search(str(value)).group(0)

    return value

#Creates spreadsheet for data storage
def sheetCreate():
    print("--Creating tfArbitrage.xlsx in current directory")
    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Scrap.tf ID"
    sheet["B1"] = "Name"
    sheet["C1"] = "Item Quality"
    sheet["D1"] = "Quantity"
    sheet["E1"] = "Price (Total in Refined)"
    sheet["F1"] = "BP.tf Price (Total in Refined)"
    sheet["G1"] = "PROFIT"

    workbook.save(filename="tfArbitrage.xlsx")


#CURRENTLY UNIMPLEMENTED
#Will create listings for each item considered "Viable" on the spreadsheet
def listingPosts():
    print("-- Posting listings for inventory.")
    tradeBot()

#Will respond to appropriate messages on steam
    #Should be running consistently, as the rest of the program is running. Or run itermittently as it checks for messages.
def tradeBot():
    print("-- Managing incoming trade offers and messages.")
    listingPosts()


def main():
    if(os.path.exists(sheetDirectory)) != True:
        sheetCreate()
    authCode = getAuthCode()
    scrapeScrap(authCode)

main()
